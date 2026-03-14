import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
import random

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Sales Data"

# Add headers
headers = ["ID", "Date", "Product", "Category", "Quantity", "Unit Price", "Total", "Sales Rep", "Region"]
ws.append(headers)

# Sample data
products = ["Laptop", "Mouse", "Keyboard", "Monitor", "Headphones", "Webcam", "USB Cable", "Desk Lamp"]
categories = ["Electronics", "Accessories", "Peripherals", "Office Supplies"]
sales_reps = ["John Smith", "Emma Davis", "Michael Brown", "Sarah Wilson", "David Lee"]
regions = ["North", "South", "East", "West", "Central"]

# Generate 50 rows of dummy data
start_date = datetime(2024, 1, 1)

for i in range(1, 51):
    row_id = i
    date = start_date + timedelta(days=random.randint(0, 365))
    product = random.choice(products)
    category = random.choice(categories)
    quantity = random.randint(1, 20)
    unit_price = round(random.uniform(10, 500), 2)
    total = round(quantity * unit_price, 2)
    sales_rep = random.choice(sales_reps)
    region = random.choice(regions)
    
    ws.append([row_id, date.strftime("%Y-%m-%d"), product, category, quantity, unit_price, total, sales_rep, region])

# Format the header row
for cell in ws[1]:
    cell.font = openpyxl.styles.Font(bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
wb.save("sales_data.xlsx")
print("Excel file 'sales_data.xlsx' created successfully with 50 rows of dummy data!")
